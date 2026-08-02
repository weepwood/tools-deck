#!/usr/bin/env python3
import json
import os
import sys
import tempfile
from pathlib import Path

PREFIX = "::tools-deck::"


def emit(payload):
    print(PREFIX + json.dumps(payload, ensure_ascii=False), flush=True)


def main():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl。请运行：python -m pip install openpyxl") from exc

    params = json.loads(os.environ.get("TOOLS_DECK_PARAMS_JSON", "{}"))
    raw_files = params.get("files", [])
    if isinstance(raw_files, str):
        files = [Path(value.strip()) for value in raw_files.replace(";", "\n").splitlines() if value.strip()]
    else:
        files = [Path(str(value)) for value in raw_files]
    if not files:
        raise ValueError("至少选择一个 Excel 文件")

    sheet_name = str(params.get("sheet", "Sheet1"))
    add_source = bool(params.get("sourceColumn", True))
    output_name = str(params.get("outputName", "merged.xlsx"))
    output_dir = Path(tempfile.gettempdir()) / "tools-deck" / os.environ.get("TOOLS_DECK_RUN_ID", "run")
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(output_name).name
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"
    output_path = output_dir / safe_name

    target = Workbook()
    target_sheet = target.active
    target_sheet.title = "Merged"
    reference_header = None
    total_rows = 0
    merged_files = 0

    for index, file in enumerate(files, start=1):
        if not file.is_file():
            raise ValueError(f"文件不存在：{file}")

        workbook = load_workbook(file, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"{file.name} 中不存在工作表 {sheet_name}")
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                emit({"type": "progress", "progress": min(95, 5 + round(index / len(files) * 90)), "message": f"[{index}/{len(files)}] 跳过空文件 {file.name}", "level": "warning"})
                continue

            normalized_header = tuple("" if value is None else str(value).strip() for value in header)
            if reference_header is None:
                reference_header = normalized_header
                target_sheet.append(list(header) + (["来源文件"] if add_source else []))
            elif normalized_header != reference_header:
                raise ValueError(f"{file.name} 的表头与第一个有效文件不一致")

            for row in rows:
                target_sheet.append(list(row) + ([file.name] if add_source else []))
                total_rows += 1
            merged_files += 1
        finally:
            workbook.close()

        emit({"type": "progress", "progress": min(95, 5 + round(index / len(files) * 90)), "message": f"[{index}/{len(files)}] 已合并 {file.name}"})

    if reference_header is None:
        raise ValueError("所选文件中没有可合并的数据")

    target.save(output_path)
    emit({"type": "artifact", "progress": 100, "artifact": {"type": "file", "label": output_path.name, "path": str(output_path), "content": str(output_path)}})
    print(f"合并完成：{merged_files} 个有效文件，{total_rows} 行", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001
        print(f"Excel 合并失败：{exc}", file=sys.stderr, flush=True)
        raise
