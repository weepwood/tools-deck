import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image

BUILTIN_DIR = Path(__file__).resolve().parent.parent


def run_script(name, params, run_id):
    env = os.environ.copy()
    env["TOOLS_DECK_PARAMS_JSON"] = json.dumps(params, ensure_ascii=False)
    env["TOOLS_DECK_RUN_ID"] = run_id
    return subprocess.run(
        [sys.executable, str(BUILTIN_DIR / name)],
        env=env,
        text=True,
        capture_output=True,
        check=False,
    )


class ImageCompressorTests(unittest.TestCase):
    def test_rejects_output_inside_source(self):
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "source"
            source.mkdir()
            output = source / "output"
            result = run_script(
                "image-compressor.py",
                {"input": str(source), "output": str(output)},
                "image-invalid-path",
            )
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("不能位于输入文件夹内部", result.stderr)

    def test_compresses_image_to_separate_directory(self):
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "source"
            output = Path(temporary) / "output"
            source.mkdir()
            Image.new("RGB", (8, 8), "white").save(source / "sample.jpg")
            result = run_script(
                "image-compressor.py",
                {"input": str(source), "output": str(output), "quality": 80},
                "image-success",
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue((output / "sample.jpg").is_file())


class ExcelMergerTests(unittest.TestCase):
    def create_workbook(self, path, header, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(header)
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    def test_merges_matching_headers(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            self.create_workbook(first, ["id", "name"], [[1, "A"]])
            self.create_workbook(second, ["id", "name"], [[2, "B"]])
            run_id = "excel-success"
            result = run_script(
                "excel-merger.py",
                {"files": [str(first), str(second)], "outputName": "merged.xlsx"},
                run_id,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            output = Path(tempfile.gettempdir()) / "tools-deck" / run_id / "merged.xlsx"
            workbook = load_workbook(output, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
            self.assertEqual(rows[0], ("id", "name", "来源文件"))
            self.assertEqual(len(rows), 3)

    def test_rejects_mismatched_headers(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            self.create_workbook(first, ["id", "name"], [[1, "A"]])
            self.create_workbook(second, ["id", "title"], [[2, "B"]])
            result = run_script(
                "excel-merger.py",
                {"files": [str(first), str(second)]},
                "excel-invalid-header",
            )
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("表头与第一个有效文件不一致", result.stderr)


if __name__ == "__main__":
    unittest.main()
